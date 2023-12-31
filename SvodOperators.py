import datetime
import shutil
import csv
import openpyxl
from progress.spinner import Spinner
from operator import itemgetter
import numpy as np
import os

# просто крутилка в консоли
spinner = Spinner('Обработка данных:')


def data_filter_max_speed(max_speed_data):
    # поиск макисмальной скорости у населенного пункта

    max_speed_for_sort = []

    max_speed = ''
    count = 0
    is_convert = False

    for i in max_speed_data:
        max_speed_split = i.split(' ')

        # перевод мегабит в килобит для сортировки
        if max_speed_split[1] == 'Мб/с':
            mb_to_kb = int(max_speed_split[0]) * 1000
            max_speed_split[0] = mb_to_kb
            max_speed_split[1] = 'Кб/с'
            is_convert = True

        max_speed_for_sort.append(
            [int(max_speed_split[0]), max_speed_split[1]])

        count += 1

        if count == len(max_speed_data):
            count = 0
            break

    max_speed_for_sort.sort(key=lambda x: x[0], reverse=True)

    if is_convert:
        # перевод килобит в мегабит для вывода
        speed = round(max_speed_for_sort[0][0] / 1000)
        max_speed = str(speed) + ' ' + 'Мб/с'
    else:
        max_speed = str(max_speed_for_sort[0][0]) + \
            ' ' + max_speed_for_sort[0][1]

    return max_speed


def data_sum_columns(columns_for_sum):
    # суммирование столбцов по вертикали
    sum = np.sum(columns_for_sum, axis=0)
    return sum


def write_last_row(i, check_new_fias, work_sheet, max_speed, columns_for_sum, is_ts, is_uslugi_svyazi, is_gsm, is_umts, is_lte, count_row):
    # запись последней строки в файле
    # сделано пока так, потому что потому
    ts_str = 'нет'
    uslugi_svyazi_str = 'нет'
    gsm_str = 'нет'
    umts_str = 'нет'
    lte_str = 'нет'

    if is_ts:
        ts_str = 'да'
    if is_uslugi_svyazi:
        uslugi_svyazi_str = 'да'
    if is_gsm:
        gsm_str = 'да'
    if is_umts:
        umts_str = 'да'
    if is_lte:
        lte_str = 'да'

    is_ts = False
    is_uslugi_svyazi = False
    is_gsm = False
    is_umts = False
    is_lte = False

    work_sheet['W' + str(count_row)] = ts_str  # ТС
    # Услуги связи
    work_sheet['X' + str(count_row)] = uslugi_svyazi_str
    work_sheet['T' + str(count_row)] = gsm_str  # GSM
    work_sheet['U' + str(count_row)] = umts_str  # UMTS
    work_sheet['V' + str(count_row)] = lte_str  # LTE

    for row in work_sheet.iter_rows(min_row=count_row, max_row=count_row, min_col=8, max_col=19):
        for cell in row:
            # print(cell.value)
            if cell.value is None:
                cell.value = 'нет'

    # этот говнокод ничто иное, как суммарная информация по всем операторам
    svod_mts = 'МТС - GSM - '+str(work_sheet['H' + str(count_row)].value)+'; МТС - UMTS - '+str(
        work_sheet['I' + str(count_row)].value)+'; МТС - LTE - '+str(work_sheet['J' + str(count_row)].value)+'; '
    svod_megafon = 'МегаФон - GSM - '+str(work_sheet['K' + str(count_row)].value)+'; МегаФон - UMTS - '+str(
        work_sheet['L' + str(count_row)].value)+'; МегаФон - LTE - '+str(work_sheet['M' + str(count_row)].value)+'; '
    svod_beeline = 'Билайн - GSM - '+str(work_sheet['N' + str(count_row)].value)+'; Билайн - UMTS - '+str(
        work_sheet['O' + str(count_row)].value)+'; Билайн - LTE - '+str(work_sheet['P' + str(count_row)].value)+'; '
    svod_tele2 = 'Теле2 - GSM - '+str(work_sheet['Q' + str(count_row)].value)+'; Теле2 - UMTS - '+str(
        work_sheet['R' + str(count_row)].value)+'; Теле2 - LTE - '+str(work_sheet['S' + str(count_row)].value)+';'

    work_sheet['AD' + str(count_row)] = svod_mts + \
        svod_megafon + svod_beeline + svod_tele2

    # замена ячеек с циферами на 'да'
    if work_sheet['H' + str(count_row)].value != 'нет':
        work_sheet['H' + str(count_row)] = 'да'
    if work_sheet['K' + str(count_row)].value != 'нет':
        work_sheet['K' + str(count_row)] = 'да'
    if work_sheet['N' + str(count_row)].value != 'нет':
        work_sheet['N' + str(count_row)] = 'да'
    if work_sheet['Q' + str(count_row)].value != 'нет':
        work_sheet['Q' + str(count_row)] = 'да'

    # запись максимальной скорости
    work_sheet['Y' + str(count_row)] = data_filter_max_speed(max_speed)
    max_speed.clear()

    sum = data_sum_columns(columns_for_sum)
    work_sheet['Z' + str(count_row)] = sum[0]  # кол-во каналов
    work_sheet['AA' + str(count_row)] = sum[1]  # кол-во таксофонов
    work_sheet['AB' + str(count_row)] = sum[2]  # кол-во точек доступа
    columns_for_sum.clear()

    check_new_fias.append(i[1])

############################################################################################################


def data_writer(work_sheet, data):
    count = 0
    count_row = 2

    check_new_fias = []
    max_speed = []
    columns_for_sum = []

    is_ts = False
    is_uslugi_svyazi = False
    is_gsm = False
    is_umts = False
    is_lte = False

    for i in data:
        spinner.next()

        # region новый населенный пункт
        if count == 0:
            check_new_fias.append(i[1])

        if i[1] not in check_new_fias:
            # запись новой строки (новый населенный пункт)
            ts_str = 'нет'
            uslugi_svyazi_str = 'нет'
            gsm_str = 'нет'
            umts_str = 'нет'
            lte_str = 'нет'

            if is_ts:
                ts_str = 'да'
            if is_uslugi_svyazi:
                uslugi_svyazi_str = 'да'
            if is_gsm:
                gsm_str = 'да'
            if is_umts:
                umts_str = 'да'
            if is_lte:
                lte_str = 'да'

            is_ts = False
            is_uslugi_svyazi = False
            is_gsm = False
            is_umts = False
            is_lte = False

            work_sheet['W' + str(count_row)] = ts_str  # ТС
            # Услуги связи
            work_sheet['X' + str(count_row)] = uslugi_svyazi_str
            work_sheet['T' + str(count_row)] = gsm_str  # GSM
            work_sheet['U' + str(count_row)] = umts_str  # UMTS
            work_sheet['V' + str(count_row)] = lte_str  # LTE

            for row in work_sheet.iter_rows(min_row=count_row, max_row=count_row, min_col=8, max_col=19):
                for cell in row:
                    # print(cell.value)
                    if cell.value is None:
                        cell.value = 'нет'

            # этот говнокод ничто иное, как суммарная информация по всем операторам
            svod_mts = 'МТС - GSM - '+str(work_sheet['H' + str(count_row)].value)+'; МТС - UMTS - '+str(
                work_sheet['I' + str(count_row)].value)+'; МТС - LTE - '+str(work_sheet['J' + str(count_row)].value)+'; '
            svod_megafon = 'МегаФон - GSM - '+str(work_sheet['K' + str(count_row)].value)+'; МегаФон - UMTS - '+str(
                work_sheet['L' + str(count_row)].value)+'; МегаФон - LTE - '+str(work_sheet['M' + str(count_row)].value)+'; '
            svod_beeline = 'Билайн - GSM - '+str(work_sheet['N' + str(count_row)].value)+'; Билайн - UMTS - '+str(
                work_sheet['O' + str(count_row)].value)+'; Билайн - LTE - '+str(work_sheet['P' + str(count_row)].value)+'; '
            svod_tele2 = 'Теле2 - GSM - '+str(work_sheet['Q' + str(count_row)].value)+'; Теле2 - UMTS - '+str(
                work_sheet['R' + str(count_row)].value)+'; Теле2 - LTE - '+str(work_sheet['S' + str(count_row)].value)+';'

            work_sheet['AD' + str(count_row)] = svod_mts + \
                svod_megafon + svod_beeline + svod_tele2

            # замена ячеек с циферами на 'да'
            if work_sheet['H' + str(count_row)].value != 'нет':
                work_sheet['H' + str(count_row)] = 'да'
            if work_sheet['K' + str(count_row)].value != 'нет':
                work_sheet['K' + str(count_row)] = 'да'
            if work_sheet['N' + str(count_row)].value != 'нет':
                work_sheet['N' + str(count_row)] = 'да'
            if work_sheet['Q' + str(count_row)].value != 'нет':
                work_sheet['Q' + str(count_row)] = 'да'

            # запись максимальной скорости
            work_sheet['Y' + str(count_row)] = data_filter_max_speed(max_speed)
            max_speed.clear()

            sum = data_sum_columns(columns_for_sum)
            work_sheet['Z' + str(count_row)] = sum[0]  # кол-во каналов
            work_sheet['AA' + str(count_row)] = sum[1]  # кол-во таксофонов
            work_sheet['AB' + str(count_row)] = sum[2]  # кол-во точек доступа
            columns_for_sum.clear()

            check_new_fias.append(i[1])
            count_row += 1
        # endregion

        max_speed.append(i[10])
        columns_for_sum.append([int(i[14]), int(i[15]), int(i[16])])

        work_sheet['A' + str(count_row)] = i[0]  # ID
        work_sheet['B' + str(count_row)] = i[1]  # ФИАС
        work_sheet['C' + str(count_row)] = i[2]  # код региона
        work_sheet['D' + str(count_row)] = i[3]  # наименование региона
        work_sheet['E' + str(count_row)] = i[4]  # город
        work_sheet['F' + str(count_row)] = i[5]  # район
        work_sheet['G' + str(count_row)] = i[6]  # населенный пункт
        work_sheet['AC' + str(count_row)] = i[17]  # url для проверки'

        # запись операторов мтс, мегафон, билайн, теле2
        if 'МТС' in i[7] or 'мтс' in i[7]:
            if work_sheet['H' + str(count_row)].value == 'нет' or work_sheet['H' + str(count_row)].value is None:
                work_sheet['H' + str(count_row)] = i[11]
            if work_sheet['I' + str(count_row)].value == 'нет' or work_sheet['I' + str(count_row)].value is None:
                work_sheet['I' + str(count_row)] = i[12]
            if work_sheet['J' + str(count_row)].value == 'нет' or work_sheet['J' + str(count_row)].value is None:
                work_sheet['J' + str(count_row)] = i[13]
        if 'МегаФон' in i[7]:
            if work_sheet['K' + str(count_row)].value == 'нет' or work_sheet['K' + str(count_row)].value is None:
                work_sheet['K' + str(count_row)] = i[11]
            if work_sheet['L' + str(count_row)].value == 'нет' or work_sheet['L' + str(count_row)].value is None:
                work_sheet['L' + str(count_row)] = i[12]
            if work_sheet['M' + str(count_row)].value == 'нет' or work_sheet['M' + str(count_row)].value is None:
                work_sheet['M' + str(count_row)] = i[13]
        if 'Вымпелк' in i[7]:
            if work_sheet['N' + str(count_row)].value == 'нет' or work_sheet['N' + str(count_row)].value is None:
                work_sheet['N' + str(count_row)] = i[11]
            if work_sheet['O' + str(count_row)].value == 'нет' or work_sheet['O' + str(count_row)].value is None:
                work_sheet['O' + str(count_row)] = i[12]
            if work_sheet['P' + str(count_row)].value == 'нет' or work_sheet['P' + str(count_row)].value is None:
                work_sheet['P' + str(count_row)] = i[13]
        if 'Т2' in i[7]:
            if work_sheet['Q' + str(count_row)].value == 'нет' or work_sheet['Q' + str(count_row)].value is None:
                work_sheet['Q' + str(count_row)] = i[11]
            if work_sheet['R' + str(count_row)].value == 'нет' or work_sheet['R' + str(count_row)].value is None:
                work_sheet['R' + str(count_row)] = i[12]
            if work_sheet['S' + str(count_row)].value == 'нет' or work_sheet['S' + str(count_row)].value is None:
                work_sheet['S' + str(count_row)] = i[13]

        if i[8] == 'да':
            is_ts = True
        if i[9] == 'да':
            is_uslugi_svyazi = True
        # gsm
        if i[11] == '900' or i[11] == '1800' or i[11] == '900/1800':
            is_gsm = True
        if i[12] == 'да':
            is_umts = True
        if i[13] == 'да':
            is_lte = True

        count += 1

        # запись последней строки (иначе не пишет данные по операторам и т.д.)
        if count == len(data):
            write_last_row(i, check_new_fias, work_sheet, max_speed,
                           columns_for_sum, is_ts, is_uslugi_svyazi, is_gsm, is_umts, is_lte, count_row)

    spinner.finish()
    return

############################################################################################################


def write_xlsx_file(data):

    work_book = openpyxl.Workbook()
    work_sheet = work_book.active
    work_sheet.title = "Свод по операторам"

    # region Заполняем шапку таблицы
    work_sheet['A1'] = 'Уникальный идентификатор населенного пункта в информационной системе Роскомнадзора'
    work_sheet['B1'] = 'Код ФИАС'
    work_sheet['C1'] = 'Код региона'
    work_sheet['D1'] = 'Наименование региона'
    work_sheet['E1'] = 'Город'
    work_sheet['F1'] = 'Район'
    work_sheet['G1'] = 'Населенный пункт'
    work_sheet['H1'] = 'МТС GSM'
    work_sheet['I1'] = 'МТС UMTS'
    work_sheet['J1'] = 'МТС LTE'
    work_sheet['K1'] = 'МегаФон GSM'
    work_sheet['L1'] = 'МегаФон UMTS'
    work_sheet['M1'] = 'МегаФон LTE'
    work_sheet['N1'] = 'Билайн GSM'
    work_sheet['O1'] = 'Билайн UMTS'
    work_sheet['P1'] = 'Билайн LTE'
    work_sheet['Q1'] = 'Теле2 GSM'
    work_sheet['R1'] = 'Теле2 UMTS'
    work_sheet['S1'] = 'Теле2 LTE'
    work_sheet['T1'] = 'Наличие GSM'
    work_sheet['U1'] = 'Наличие UMTS'
    work_sheet['V1'] = 'Наличие LTE'
    work_sheet['W1'] = 'Услуги местной ТС'
    work_sheet['X1'] = 'Телематические услуги связи'
    # телематические услуги связи (скорость) и единицы измероение в одну ячейку
    work_sheet['Y1'] = 'Телематические услуги связи (максимальная скорость передачи данных)'
    # сумма
    work_sheet['Z1'] = 'Эфирное телевидение (количество цифровых каналов)'
    work_sheet['AA1'] = 'Количество таксофонов'  # сумма
    work_sheet['AB1'] = 'Количество точек доступа'  # сумма
    work_sheet['AC1'] = 'URL'  # url для проверки
    work_sheet['AD1'] = 'Доп. информация'

    # дополнительная информация в виде:
    # 4 основных оператора:
    # МТС- GSM-900/1800; МТС- UMTS- нет; МТС- LTE - нет;
    # Билайн - GSM-900; Билайн - UMTS- нет; Билайн - LTE - нет;
    # Мегафон - GSM-нет; Мегафон - UMTS- нет; Мегафон - LTE - нет;
    # Теле2 - GSM-нет; Теле2 - UMTS- нет; Теле2 - LTE - нет
    # endregion

    data_writer(work_sheet, data)

    # сохранение получившейся какахи
    try:
        work_book.save('./processed_file/%s' % 'Свод по операторам.xlsx')
    except PermissionError:
        print(
            "Файл открыт в другой программе. Закройте файл и повторите попытку")
        return

############################################################################################################


def read_csv_file(file_name):
    data = []
    count = 0

    # file_data_sort = open('data_sort.txt', 'w')

    try:
        with open(file_name, 'r', newline='', encoding='Windows-1251') as r_file:
            reader = csv.reader(r_file, delimiter=';')
            # добавляем в список город, операторов и наличие типов связи
            for row in reader:
                if count == 0:
                    count += 1
                    continue
                # объединяем показания скорости и единицы измерения
                merge_rows = row[11] + ' ' + row[12]
                # все необходимые данные
                data.append([row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10],
                             merge_rows, row[13], row[14], row[15], row[16], row[17], row[18], row[19]])

            # сортируем по населенному пункту и району (фиас для контроля, но и без него заэбись)
            data = sorted(data, key=itemgetter(6, 5, 1))

            # n = []
            # for i in data:
            #     # if i[1] not in n:
            #     #     n.append(i[1])
            #     file_data_sort.write(str(i[1]) + ' - ' + str(i[3]) + ' - ' + str(i[4]) + ' - ' + str(
            #         i[5]) + ' - ' + str(i[6]) + ' -( ' + str(i[8]) + ' - ' + str(i[9]) + ' )-{ ' + ' - ' + str(i[11]) + ' - ' + str(i[12]) + ' - ' + str(i[13]) + ' - ' + str(i[7]) + '}\n')
            # file_data_sort.close()

    except IOError:
        print(
            "Файл открыт в другой программе. Закройте файл и повторите попытку")
        return

    return data

############################################################################################################


def main():
    print('###############################################')
    print('Начало обработки файла Выгрузка.csv')

    bd = datetime.datetime.now()

    csv_file = './processed_file/%s' % 'Выгрузка.csv'
    csv_file_dest = './source_file/%s' % 'Выгрузка.csv'

    try:
        shutil.copyfile(csv_file, csv_file_dest)
    except IOError:
        print("Файл Выгрузка.csv не найден")
        return

    write_xlsx_file(read_csv_file(csv_file_dest))

    os.remove(csv_file_dest)

    ed = datetime.datetime.now()
    total_time = ed - bd
    print('Обработка файла Выгрузка.csv завершена')
    print('Файл Свод по операторам.xlsx сохранен в папке processed_file')
    print('Время выполения программы: ')
    print(total_time)
    return


main()
print('Нажмите Enter для выхода')
input()
