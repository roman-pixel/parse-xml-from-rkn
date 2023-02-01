import os
import fnmatch
import csv
import openpyxl
import datetime
from progress.spinner import Spinner
import shutil


def read_xlsx_file(file):
    arr_id_url = []

    # workbook object is created
    wb_obj = openpyxl.load_workbook(file)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    spinner = Spinner('Обработка Excel файла ')

    # чтение файла со 2 строки
    # запись 2 и 13 столбцов в массив (id и URL)
    for i in range(2, m_row + 1):
        cell_id = sheet_obj.cell(row=i, column=2)
        cell_url = sheet_obj.cell(row=i, column=13)
        arr_id_url.append([cell_id.value, cell_url.value])
        spinner.next()
    spinner.finish()
    print('Выполнено')
    return arr_id_url

###############################################


def edit_csv_file(csv_file, arr_id_url):
    try:
        with open(csv_file, encoding='Windows-1251') as r_file:
            # Создаем объект reader, указываем символ-разделитель ";"
            file_reader = csv.reader(r_file, delimiter=";")
            itogi_new = []
            count = 0
            count_find_iter = 0
            count_huanter = 0

            spinner = Spinner('Обработка CSV файла ')
                      
            for row in file_reader:
                spinner.next()
                # запись названия столбца
                if count == 0:
                    row.insert(1, 'ID')
                    itogi_new.append(row)
                    count += 1
                    continue

                for j in arr_id_url:
                    if row[18] == j[1]:
                        row.insert(1, j[0])
                        itogi_new.append(row)
                        count_find_iter = 0
                        count_huanter += 1
                        break

                    count_find_iter += 1

                    if count_find_iter == len(arr_id_url):
                        row.insert(1, '')
                        itogi_new.append(row)
                        count_find_iter = 0

            spinner.finish()
            print ('Найдено %s записей' % count_huanter)
            print('Выполнено')
    except IOError:
        print("Файл itog.csv не найден")
        return

    ##########################

    # запись в файл
    try:
        with open('./processed_file/%s' % 'itog_with_id.csv', 'w', newline='', encoding='Windows-1251') as w_file:
            writer = csv.writer(w_file, delimiter=';')
            writer.writerows(itogi_new)
    except IOError:
        print("Файл открыт в другой программе. Закройте файл и повторите попытку.")

    r_file.close()
    w_file.close()
    os.remove(csv_file)
    return

###############################################


def main():
    bd = datetime.datetime.now()

    csv_file_src = './processed_file/%s' % 'itog.csv'
    csv_file_dest = './source_file/%s' % 'itog.csv'

    try:
        shutil.copyfile(csv_file_src, csv_file_dest)
    except IOError:
        print("Файл itog.csv не найден")
        return

    # ищем тот самый файл с расширением xml
    xlsx_file = ''
    count_of_xlsx_files = 0
    listOfFiles = os.listdir('./source_file')
    pattern = "*.xlsx"
    for entry in listOfFiles:
        if fnmatch.fnmatch(entry, pattern):
            count_of_xlsx_files += 1

            if count_of_xlsx_files > 1:
                print('Найдено более одного файла с расширением xlsx')
                return

            xlsx_file = entry

    if xlsx_file == '':
        print('Файл с расширением xlsx не найден')
        return

    edit_csv_file(csv_file_dest, read_xlsx_file('./source_file/%s' % xlsx_file))
    ed = datetime.datetime.now()
    total_time = ed - bd
    print('Время выполения программы: ')
    print(total_time)
    return

###############################################


main()
input('Нажмите Enter для выхода.')
