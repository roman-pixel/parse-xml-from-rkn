import os
import fnmatch
import csv
import openpyxl
from progress.spinner import Spinner


def add_fias_excel(file, data):
    # workbook object is created
    wb_obj = openpyxl.load_workbook(file)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    spinner = Spinner('Обработка Excel файла ')

    sheet_obj.cell(row=1, column=31, value='Код ФИАС')

    for i in range(2, m_row + 1):
        cell_url = sheet_obj.cell(row=i, column=13)

        for j in data:
            if cell_url.value == j[1]:
                sheet_obj.cell(row=i, column=31, value=j[0])
                break

        spinner.next()
    spinner.finish()
    print('Выполнено')
    wb_obj.save('./processed_file/%s' % 'with_fias.xlsx')


def read_csv_file(csv_file):
    data = []
    count_iter = 0

    with open(csv_file, encoding='Windows-1251') as r_file:
        # Создаем объект reader, указываем символ-разделитель ","
        file_reader = csv.reader(r_file, delimiter=";")
        for row in file_reader:
            if count_iter != 0:
                # print(row[2], row[18])
                data.append([row[2], row[18]])
            count_iter += 1
    return data


def main():
    csv_file_src = './processed_file/%s' % 'itog.csv'

    # ищем тот самый файл с расширением xml
    xlsx_file = ''
    count_of_xlsx_files = 0
    listOfFiles = os.listdir('./source_file')
    pattern = "*.xlsx"
    for entry in listOfFiles:
        if fnmatch.fnmatch(entry, pattern):
            count_of_xlsx_files += 1

            if count_of_xlsx_files > 1:
                print('Найдено более одного файла с расширением xlsx')
                return

            xlsx_file = entry

    if xlsx_file == '':
        print('Файл с расширением xlsx не найден')
        return

    add_fias_excel('./source_file/%s' % xlsx_file, read_csv_file(csv_file_src))


main()
